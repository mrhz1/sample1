import os
import re
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pydicom
from openpyxl import Workbook, load_workbook

DATE_RE = re.compile(r"(\d{2})[-_.](\d{2})[-_.](\d{4})")

# How often, in files examined, to emit a heartbeat line while scanning a
# single R-code folder. Large folders otherwise look hung for hours.
PROGRESS_EVERY = 2000

SUMMARY_HEADER = [
    "Code",
    "Study Date",
    "Modality",
    "DICOM File Count",
    "Matched PDF File(s)",
    "Match Status",
]
SUMMARY_WIDTHS = (12, 14, 10, 16, 45, 40)

# Characters that can't go in a filename, in case an R-code folder has one.
UNSAFE_NAME_RE = re.compile(r'[\\/:*?"<>|]')

# PDF-filename token (uppercased) -> standard DICOM Modality code.
# Extend this as you find more variants in the real filenames.
MODALITY_ALIASES = {
    "CT": "CT",
    "MRI": "MR",
    "MR": "MR",
    "ECHO": "US",
    "ECHOCARDIOGRAM": "US",
    "ULTRASOUND": "US",
    "US": "US",
    "XRAY": "CR",
    "X-RAY": "CR",
    "DX": "DX",
    "MAMMO": "MG",
    "MG": "MG",
    "PET": "PT",
    "NM": "NM",
}
MODALITY_RE = re.compile(
    r"(?<![A-Z0-9])("
    + "|".join(re.escape(k) for k in MODALITY_ALIASES)
    + r")(?![A-Z0-9])"
)

# Raw DICOM Modality tag value (uppercased) -> human-readable label, for
# display in the Excel outputs only. Real-world files don't always use the
# standard 2-letter code (e.g. some write "Ultrasonic" instead of "US") -
# extend this as you encounter more raw values in the real data.
MODALITY_DISPLAY_NAMES = {
    "CT": "CT",
    "MR": "MRI",
    "US": "Echo",
    "ULTRASOUND": "Echo",
    "ULTRASONIC": "Echo",
    "CR": "X-Ray",
    "DX": "X-Ray",
    "MG": "Mammography",
    "PT": "PET",
    "NM": "Nuclear Medicine",
    "XA": "Angiography",
    "RF": "Fluoroscopy",
    "SR": "Structured Report",
}

_log_fh = None


def open_log(path):
    """Append-mode log so successive (resumed) runs stack up in one place."""
    global _log_fh
    _log_fh = open(path, "a", encoding="utf-8")
    _log_fh.write(f"\n===== run started {datetime.now():%Y-%m-%d %H:%M:%S} =====\n")
    _log_fh.flush()


def log(message):
    line = f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {message}"
    print(line, flush=True)
    if _log_fh is not None:
        _log_fh.write(line + "\n")
        _log_fh.flush()


def format_duration(seconds):
    seconds = int(seconds)
    hours, rest = divmod(seconds, 3600)
    minutes, secs = divmod(rest, 60)
    if hours:
        return f"{hours}h{minutes:02d}m{secs:02d}s"
    if minutes:
        return f"{minutes}m{secs:02d}s"
    return f"{secs}s"


def display_modality(code):
    if not code:
        return code
    return MODALITY_DISPLAY_NAMES.get(code.strip().upper(), code)


def result_path(results_dir, rcode):
    return results_dir / f"{UNSAFE_NAME_RE.sub('_', rcode)}-results.xlsx"


def parse_pdf_date(name):
    match = DATE_RE.search(name)
    if not match:
        return None
    month, day, year = match.groups()
    return f"{year}{month}{day}"


def parse_pdf_modality(name):
    match = MODALITY_RE.search(name.upper())
    if not match:
        return None
    return MODALITY_ALIASES[match.group(1)]


def read_dicom_study(path, debug=False):
    try:
        ds = pydicom.dcmread(str(path), stop_before_pixels=True, force=False)
    except Exception as exc:
        if debug:
            forced_ok = False
            forced_date = None
            try:
                ds2 = pydicom.dcmread(str(path), stop_before_pixels=True, force=True)
                forced_ok = True
                forced_date = str(ds2.get("StudyDate", ""))
            except Exception:
                pass
            hint = ""
            if forced_ok:
                hint = (
                    f" -> READABLE with force=True (StudyDate={forced_date!r}); "
                    "this file is likely missing the DICOM preamble/DICM header"
                )
            print(f"  [SKIP] {path}: {type(exc).__name__}: {exc}{hint}")
        return None
    study_date, modality = str(ds.get("StudyDate", "")), str(ds.get("Modality", ""))
    if debug:
        print(f"  [OK]   {path}: StudyDate={study_date!r} Modality={modality!r}")
    return study_date, modality


def scan_reports_dir(rcode_dir, debug=False):
    """One R-code's reports folder -> list of {"path", "date", "modality"}."""
    entries = []
    for path in sorted(rcode_dir.rglob("*.pdf")):
        if not path.is_file():
            continue
        date = parse_pdf_date(path.name)
        modality = parse_pdf_modality(path.name)
        entries.append({"path": path, "date": date, "modality": modality})
        if debug:
            print(f"  {path.name!r} -> date={date!r} modality={modality!r}")
    return entries


def build_indices(rcode_reports):
    """entries -> (by_date_modality, by_date) lookup dicts for one R-code."""
    by_date_modality = defaultdict(list)
    by_date = defaultdict(list)
    for entry in rcode_reports:
        if entry["date"] is None:
            continue
        by_date[entry["date"]].append(entry)
        if entry["modality"] is not None:
            by_date_modality[(entry["date"], entry["modality"])].append(entry)
    return by_date_modality, by_date


def process_rcode(rcode, images_dir, reports_dir, debug=False):
    """Scan one R-code end to end. Returns (rows, stats)."""
    started = time.time()
    rcode_reports = scan_reports_dir(reports_dir, debug=debug) if reports_dir else []
    by_date_modality, by_date = build_indices(rcode_reports)

    # (study_date, modality) -> {"count", "matched_pdfs"}
    summary = defaultdict(lambda: {"count": 0, "matched_pdfs": set()})
    matched_pdf_paths = set()
    files_seen = dicom_seen = 0

    if images_dir is not None:
        for path in images_dir.rglob("*"):
            if not path.is_file():
                continue
            files_seen += 1
            if files_seen % PROGRESS_EVERY == 0:
                elapsed = time.time() - started
                rate = files_seen / elapsed if elapsed else 0
                log(
                    f"    {rcode}: {files_seen:,} files examined, "
                    f"{dicom_seen:,} DICOM, {format_duration(elapsed)} elapsed "
                    f"({rate:,.0f} files/s)"
                )
            info = read_dicom_study(path, debug=debug)
            if info is None:
                continue
            dicom_seen += 1
            study_date, modality = info

            if modality and (study_date, modality) in by_date_modality:
                matches = by_date_modality[(study_date, modality)]
            elif study_date in by_date:
                matches = by_date[study_date]
            else:
                matches = []

            bucket = summary[(study_date, modality)]
            bucket["count"] += 1
            for m in matches:
                matched_pdf_paths.add(m["path"])
                bucket["matched_pdfs"].add(str(m["path"]))

    rows = []
    matched = ambiguous = unmatched_images = 0
    for (study_date, modality), info in sorted(summary.items()):
        if info["matched_pdfs"]:
            status = "Matched"
            matched += 1
        elif rcode_reports:
            status = "Same folder, no date/modality match - review manually"
            ambiguous += 1
        else:
            status = "No PDF report found"
            unmatched_images += 1
        rows.append(
            [
                rcode,
                study_date,
                display_modality(modality),
                info["count"],
                ", ".join(sorted(info["matched_pdfs"])),
                status,
            ]
        )

    unmatched_pdf = 0
    for entry in sorted(
        rcode_reports, key=lambda e: (e["date"] or "", e["modality"] or "")
    ):
        if entry["path"] not in matched_pdf_paths:
            rows.append(
                [
                    rcode,
                    entry["date"] or "",
                    display_modality(entry["modality"]) or "",
                    "",
                    str(entry["path"]),
                    "No DICOM images found",
                ]
            )
            unmatched_pdf += 1

    stats = {
        "files_seen": files_seen,
        "dicom_seen": dicom_seen,
        "pdfs": len(rcode_reports),
        "matched": matched,
        "ambiguous": ambiguous,
        "unmatched_images": unmatched_images,
        "unmatched_pdf": unmatched_pdf,
        "elapsed": time.time() - started,
    }
    return rows, stats


def write_summary_sheet(path, rows, title="Summary"):
    """Write rows to a fresh workbook via a temp file, then swap it in.

    The swap matters: a half-written results file would otherwise look
    "done" to the next resumed run and get skipped.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(SUMMARY_HEADER)
    for row in rows:
        ws.append(row)
    for column, width in zip("ABCDEF", SUMMARY_WIDTHS):
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp)
    os.replace(tmp, path)


def read_result_rows(path):
    """Read back one per-R-code results file, minus its header row."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
    finally:
        wb.close()


def combine_results(results_dir, output_xlsx):
    """Rebuild the all-codes summary from every per-R-code file on disk."""
    rows = []
    files = sorted(results_dir.glob("*-results.xlsx"))
    for path in files:
        try:
            rows.extend(read_result_rows(path))
        except Exception as exc:
            log(
                f"  WARNING: could not read {path.name} for the combined summary: {exc}"
            )
    write_summary_sheet(output_xlsx, rows)
    return len(files), len(rows)


def parse_args(argv):
    debug = "--debug" in argv
    force = "--force" in argv
    rcode_filter = None
    positional = []
    i = 0
    while i < len(argv):
        a = argv[i]
        if a in ("--debug", "--force"):
            i += 1
        elif a == "--rcode":
            if i + 1 >= len(argv):
                print(
                    "--rcode requires a value, e.g. --rcode R123 or --rcode R123,R456"
                )
                sys.exit(1)
            rcode_filter = {c.strip() for c in argv[i + 1].split(",") if c.strip()}
            i += 2
        else:
            positional.append(a)
            i += 1
    return positional, debug, force, rcode_filter


def list_rcode_dirs(root, rcode_filter):
    dirs = sorted(p for p in root.iterdir() if p.is_dir())
    if rcode_filter is not None:
        dirs = [d for d in dirs if d.name in rcode_filter]
    return dirs


def main():
    args, debug, force, rcode_filter = parse_args(sys.argv[1:])

    if len(args) < 2:
        print(__doc__)
        sys.exit(1)

    images_root = Path(args[0]).expanduser().resolve()
    reports_root = Path(args[1]).expanduser().resolve()
    if not images_root.is_dir():
        print(f"Not a folder: {images_root}")
        sys.exit(1)
    if not reports_root.is_dir():
        print(f"Not a folder: {reports_root}")
        sys.exit(1)

    output_xlsx = Path(args[2]) if len(args) > 2 else Path.cwd() / "match_report.xlsx"
    results_dir = output_xlsx.parent / "results"
    results_dir.mkdir(parents=True, exist_ok=True)
    open_log(results_dir / "progress.log")

    image_dirs = {p.name: p for p in list_rcode_dirs(images_root, rcode_filter)}
    report_dirs = {p.name: p for p in list_rcode_dirs(reports_root, rcode_filter)}
    rcodes = sorted(set(image_dirs) | set(report_dirs))

    log(f"Images:  {images_root}")
    log(f"Reports: {reports_root}")
    log(f"Results: {results_dir}")
    if rcode_filter is not None:
        log(f"Restricting scan to R-code(s): {sorted(rcode_filter)}")

    if debug:
        print("\n--- R-code comparison ---")
        print(f"In both: {sorted(set(image_dirs) & set(report_dirs))}")
        print(f"Only in images: {sorted(set(image_dirs) - set(report_dirs))}")
        print(f"Only in reports: {sorted(set(report_dirs) - set(image_dirs))}")
        for name in rcodes:
            print(f"  rcode: {name!r}")
        print("--- end R-code comparison ---\n")

    todo = [c for c in rcodes if force or not result_path(results_dir, c).exists()]
    already_done = len(rcodes) - len(todo)
    log(
        f"{len(rcodes)} R-code folder(s) in scope; {already_done} already have results, {len(todo)} to do"
    )
    if already_done and not force:
        log("(pass --force to re-scan the ones that already have a results file)")

    run_started = time.time()
    totals = defaultdict(int)
    for index, rcode in enumerate(todo, start=1):
        remaining = len(todo) - index
        log(
            f"[{index}/{len(todo)}] {rcode}: starting ({remaining} folder(s) left after this)"
        )
        rows, stats = process_rcode(
            rcode, image_dirs.get(rcode), report_dirs.get(rcode), debug=debug
        )
        out = result_path(results_dir, rcode)
        write_summary_sheet(out, rows)
        for key, value in stats.items():
            totals[key] += value
        log(
            f"[{index}/{len(todo)}] {rcode}: done in {format_duration(stats['elapsed'])} - "
            f"{stats['files_seen']:,} files ({stats['dicom_seen']:,} DICOM), "
            f"{stats['pdfs']} PDF(s), {stats['matched']} matched study bucket(s), "
            f"{stats['ambiguous']} needing review, {stats['unmatched_pdf']} unmatched PDF(s) "
            f"-> {out.name}"
        )
        elapsed = time.time() - run_started
        if remaining:
            eta = elapsed / index * remaining
            log(
                f"    run elapsed {format_duration(elapsed)}, rough ETA {format_duration(eta)} for the rest"
            )

    log("Rebuilding combined summary from results/ ...")
    file_count, row_count = combine_results(results_dir, output_xlsx)

    log("")
    log(
        f"R-codes scanned this run: {len(todo)} (skipped as already done: {already_done})"
    )
    log(
        f"Files examined: {totals['files_seen']:,} ({totals['dicom_seen']:,} readable DICOM)"
    )
    log(f"Matched study buckets: {totals['matched']}")
    log(f"Same folder, no date/modality match (needs review): {totals['ambiguous']}")
    log(f"DICOM studies with no PDF report: {totals['unmatched_images']}")
    log(f"PDF reports with no DICOM images: {totals['unmatched_pdf']}")
    log(f"Total run time: {format_duration(time.time() - run_started)}")
    log(f"Per-R-code results: {results_dir}")
    log(
        f"Combined summary ({row_count:,} rows from {file_count} R-code file(s)): {output_xlsx}"
    )


if __name__ == "__main__":
    main()
